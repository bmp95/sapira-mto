"""Escribe el gold set PRE-ANOTADO por Claude, para que Bernabe lo valide.

AVISO: esto NO es una anotacion independiente. Quien la escribe ha visto la salida del
sistema. Sirve para ahorrar tecleo, no para sustituir el criterio. Las celdas marcadas
JUICIO son las que hay que revisar de verdad.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from motor.lectura_mto import leer_mto

C, I, X = "cierta", "interpretada", "indecidible"

# fila, nombre, cant, material, calidad, medida, longitud, norma, acabado
# cada atributo: (valor, confianza)
G = [
 (1,"ESPARRAGO",40,("AC",I),("GR B7",C),('7/8"',C),("130 mm",I),("ASTM A193",C),("",C)),
 (1,"TUERCA",80,("AC",I),("GR 2H",C),('7/8"',C),("N/A",C),("ASTM A194",C),("",C)),
 (1,"ARANDELA",80,("AC",I),("",X),('7/8"',C),("N/A",C),("ASTM F436",C),("",C)),
 (2,"TORNILLO",160,("INOX",I),("A4-70",C),("M20",C),("90 mm",C),("ISO 4014",C),("",C)),
 (2,"TUERCA",160,("",X),("",X),("M20",I),("N/A",C),("ISO 4032",C),("",C)),
 (3,"TORNILLO",80,("INOX",I),("A2",C),("M12",C),("50 mm",C),("ISO 4017",C),("",C)),
 (3,"TUERCA",80,("",X),("",X),("M12",I),("N/A",C),("",X),("",C)),
 (3,"ARANDELA",80,("",X),("",X),("M12",I),("N/A",C),("",X),("",C)),
 (4,"TORNILLO",100,("AC",I),("8.8",C),("M16",C),("60 mm",C),("ISO 4017",C),("CINCADO",C)),
 (4,"TUERCA",100,("",X),("",X),("M16",I),("N/A",C),("ISO 4032",C),("CINCADO",I)),
 (4,"ARANDELA",100,("",X),("",X),("M16",I),("N/A",C),("ISO 7089",C),("CINCADO",I)),
 (5,"ESPARRAGO",24,("AC",I),("GR B7",C),('1"',C),("150 mm",I),("ASTM A193",C),("",C)),
 (5,"TUERCA",48,("AC",I),("GR 2H",C),('1"',C),("N/A",C),("ASTM A194",C),("",C)),
 (5,"ARANDELA",24,("AC",I),("",X),('1"',C),("N/A",C),("ASTM F436",C),("",C)),
 (6,"TORNILLO",60,("AC",I),("8.8",C),("M16",C),("80 mm",C),("ISO 4014",C),("CINCADO",C)),
 (6,"TUERCA",60,("",X),("",X),("M16",I),("N/A",C),("ISO 4032",C),("CINCADO",I)),
 (7,"TORNILLO",50,("INOX",I),("A4-70",C),("M12",C),("60 mm",C),("ISO 4014",C),("",C)),
 (7,"TUERCA",50,("INOX",I),("A4-80",C),("M12",C),("N/A",C),("ISO 4032",C),("",C)),
 (8,"TORNILLO",75,("AC",I),("8.8",C),("M16",C),("70 mm",C),("",X),("CINCADO",C)),
 (8,"TUERCA",75,("",X),("",X),("M16",I),("N/A",C),("",X),("CINCADO",I)),
 (8,"ARANDELA",75,("",X),("",X),("M16",I),("N/A",C),("",X),("CINCADO",I)),
 (9,"ESPARRAGO",30,("AC",I),("8.8",C),("M20",C),("200 mm",C),("DIN 975",C),("CINCADO",C)),
 (9,"TUERCA",60,("",X),("",X),("M20",I),("N/A",C),("ISO 4032",C),("CINCADO",I)),
 (9,"ARANDELA",60,("",X),("",X),("M20",I),("N/A",C),("ISO 7089",C),("CINCADO",I)),
 (10,"TORNILLO",500,("AC",I),("8.8",C),("M10",C),("40 mm",C),("ISO 4017",C),("CINCADO",C)),
 (11,"TUERCA",200,("INOX",I),("A4-80",C),("M16",C),("N/A",C),("ISO 4032",C),("",C)),
 (12,"ESPARRAGO",40,("AC",I),("GR B7",C),('3/4"',C),("110 mm",I),("ASTM A193",C),("",C)),
 (13,"TUERCA",300,("AC",I),("8.8",C),("M12",C),("N/A",C),("ISO 10511",C),("CINCADO",C)),
 (14,"ARANDELA",250,("AC",C),("",X),("M10",C),("N/A",C),("ISO 7089",C),("CINCADO",C)),
 (15,"TORNILLO",120,("AC",I),("12.9",C),("M10",C),("40 mm",C),("ISO 4762",C),("GEOMET",C)),
]

ROJO, BANDA, GRIS = "B23A2E", "F4F2F0", "888888"
AMBAR, VERDE = "FFF3CD", "E8F5E9"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "gold"
cab = ["id", "fila", "MTO: descripcion", "MTO: MATERIAL", "MTO: MEDIDA", "MTO: CANT.",
       "nombre", "conf", "cantidad", "material", "conf", "calidad", "conf", "medida", "conf",
       "longitud", "conf", "norma", "conf", "acabado", "conf", "revisar", "notas"]
ws.append(cab)
for c in range(1, len(cab) + 1):
    cel = ws.cell(row=1, column=c)
    cel.font = Font(bold=True, color="FFFFFF", size=9)
    cel.fill = PatternFill("solid", fgColor=ROJO)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

filas = {f.item: f for f in leer_mto(Path("datos/MTO_tornilleria.xlsx"))}
r = 2
vista = set()
for fila, nombre, cant, *attrs in G:
    f = filas[fila]
    orden = sum(1 for g in G[:G.index((fila, nombre, cant, *attrs))] if g[0] == fila) + 1
    ws.cell(row=r, column=1, value=f"L{fila:02d}-{orden}")
    ws.cell(row=r, column=2, value=fila)
    if fila not in vista:
        vista.add(fila)
        ws.cell(row=r, column=3, value=f.descripcion).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4, value=f.material_col)
        ws.cell(row=r, column=5, value=f.medida_col)
        ws.cell(row=r, column=6, value=f.cantidad)
        for c in range(3, 7):
            ws.cell(row=r, column=c).font = Font(color=GRIS, size=9)
    ws.cell(row=r, column=7, value=nombre)
    ws.cell(row=r, column=8, value=C)
    ws.cell(row=r, column=9, value=cant)
    juicio = 0
    for i, (val, conf) in enumerate(attrs):
        cv, cc = 10 + i * 2, 11 + i * 2
        ws.cell(row=r, column=cv, value=val)
        ws.cell(row=r, column=cc, value=conf)
        if conf in (I, X):
            juicio += 1
            for c in (cv, cc):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AMBAR)
    ws.cell(row=r, column=22, value="JUICIO" if juicio else "mecanica")
    if juicio:
        ws.cell(row=r, column=22).font = Font(bold=True, color=ROJO)
    else:
        ws.cell(row=r, column=22).fill = PatternFill("solid", fgColor=VERDE)
    r += 1
ultima = r - 1

dv = DataValidation(type="list", formula1=f'"{C},{I},{X}"', allow_blank=True)
ws.add_data_validation(dv)
for ci in [8] + [11 + i * 2 for i in range(6)]:
    col = get_column_letter(ci)
    dv.add(f"{col}2:{col}{ultima}")

for col, w in [("A", 9), ("B", 6), ("C", 58), ("D", 24), ("E", 12), ("F", 9),
               ("G", 14), ("H", 12), ("I", 9), ("V", 11), ("W", 30)]:
    ws.column_dimensions[col].width = w
for i in range(6):
    ws.column_dimensions[get_column_letter(10 + i * 2)].width = 13
    ws.column_dimensions[get_column_letter(11 + i * 2)].width = 12
ws.freeze_panes = "G2"

ruta = "datos/gold_set_PREANOTADO.xlsx"
wb.save(ruta)
amb = sum(1 for g in G for v, c in g[3:] if c in (I, X))
print(f"escrito: {ruta} | {len(G)} lineas | {amb} celdas de juicio de {len(G)*7} totales")
