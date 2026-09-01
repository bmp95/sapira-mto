"""Genera la plantilla del gold set: 15 filas del MTO x 4 huecos de linea.

NO rellena ningun valor. El numero de lineas que escriba Bernabe en cada fila ES su
decision de segmentacion, y por eso no se le da hecha.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from motor.lectura_mto import leer_mto
from pathlib import Path

ATRIBUTOS = ["nombre", "material", "calidad", "medida", "longitud", "norma", "acabado"]
HUECOS = 4

ROJO = "B23A2E"
BANDA = "F4F2F0"
GRIS = "888888"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "gold"

cab = ["id", "fila", "MTO: descripcion", "MTO: MATERIAL", "MTO: MEDIDA", "MTO: CANT.",
       "nombre", "conf", "cantidad"]
for a in ATRIBUTOS[1:]:
    cab += [a, "conf"]
cab += ["notas"]
ws.append(cab)

for c in range(1, len(cab) + 1):
    cel = ws.cell(row=1, column=c)
    cel.font = Font(bold=True, color="FFFFFF", size=9)
    cel.fill = PatternFill("solid", fgColor=ROJO)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

filas = leer_mto(Path("datos/MTO_tornilleria.xlsx"))
r = 2
for f in filas:
    for h in range(HUECOS):
        ws.cell(row=r, column=1, value=f"L{f.item:02d}-{h + 1}")
        ws.cell(row=r, column=2, value=f.item)
        if h == 0:
            ws.cell(row=r, column=3, value=f.descripcion).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=4, value=f.material_col)
            ws.cell(row=r, column=5, value=f.medida_col)
            ws.cell(row=r, column=6, value=f.cantidad)
            for c in range(3, 7):
                ws.cell(row=r, column=c).font = Font(color=GRIS, size=9)
        if h % 2 == 1:
            for c in range(1, len(cab) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BANDA)
        r += 1
    ws.cell(row=r - HUECOS, column=3).alignment = Alignment(wrap_text=True, vertical="top")

ultima = r - 1

# Desplegable de confianza en las 7 columnas "conf"
dv = DataValidation(type="list", formula1='"cierta,interpretada,indecidible"', allow_blank=True)
dv.prompt = "cierta = esta escrito en el MTO | interpretada = lo decides tu con una regla | indecidible = ni las reglas ni tu lo sabeis"
dv.promptTitle = "Cuanto te fias de esta celda"
ws.add_data_validation(dv)
COLS_CONF = [8] + [11 + i * 2 for i in range(len(ATRIBUTOS) - 1)]
for ci in COLS_CONF:
    col = get_column_letter(ci)
    dv.add(f"{col}2:{col}{ultima}")

anchos = {"A": 9, "B": 6, "C": 62}
ws.column_dimensions["A"].width = anchos["A"]
ws.column_dimensions["B"].width = anchos["B"]
ws.column_dimensions["C"].width = anchos["C"]
ws.column_dimensions["D"].width = 26
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 8
ws.column_dimensions["G"].width = 15
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 9
for i in range(len(ATRIBUTOS) - 1):
    ws.column_dimensions[get_column_letter(10 + i * 2)].width = 13
    ws.column_dimensions[get_column_letter(11 + i * 2)].width = 12
ws.column_dimensions[get_column_letter(10 + (len(ATRIBUTOS) - 1) * 2)].width = 34
ws.freeze_panes = "G2"

# Hoja de instrucciones
ins = wb.create_sheet("como se rellena")
texto = [
    ("Gold set - patron de medida del caso Sapira", True),
    ("", False),
    ("Que es esto", True),
    ("Las 15 filas del MTO, con cuatro huecos de linea cada una. Escribes UNA LINEA POR MATERIAL", False),
    ("COMPRABLE. Si una fila describe un esparrago con dos tuercas y dos arandelas, son tres lineas.", False),
    ("Si describe un esparrago solo, es una linea y dejas los otros tres huecos vacios.", False),
    ("", False),
    ("El numero de lineas que escribas ES tu decision. Por eso no viene dado.", True),
    ("", False),
    ("La columna 'conf' de cada atributo", True),
    ("   cierta        - esta escrito en el MTO, no hay discusion", False),
    ("   interpretada  - lo decides tu aplicando una regla (ej: ese 130 son milimetros)", False),
    ("   indecidible   - ni las reglas ni tu lo sabeis (ej: la dureza de una arandela sin marcar)", False),
    ("", False),
    ("Las celdas 'indecidible' NO puntuan ni a favor ni en contra. Se reportan aparte, como", False),
    ("'aqui ni el sistema ni yo lo sabemos'. Es la mitad honesta del numero.", False),
    ("", False),
    ("Valores validos", True),
    ("   nombre    TORNILLO / TUERCA / ARANDELA / VARILLA ROSCADA / ESPARRAGO", False),
    ("   material  AC / INOX / otro metal / vacio si no se puede saber", False),
    ("   calidad   8.8, 10.9, 12.9, A2, A2-70, A4-70, A4-80, 304, 316, 8, 10, 100HV..300HV,", False),
    ("             GR B7, GR 2H... o vacio", False),
    ("   medida    7/8\" o M20", False),
    ("   longitud  numero con unidad (130 mm, 5 1/2\"), o N/A en tuercas y arandelas", False),
    ("   norma     ya normalizada: ISO 4017, no DIN 933", False),
    ("   acabado   CINCADO / GALVANIZADO EN CALIENTE / PAVONADO / FOSFATADO / BICROMATADO /", False),
    ("             GEOMET / DACROMET, o vacio (vacio es un valor valido)", False),
    ("", False),
    ("Dos pasadas", True),
    ("Rellena esto entero. Guardalo. Deja pasar unas horas y hazlo OTRA VEZ en una copia limpia,", False),
    ("sin mirar la primera. El porcentaje en que no coincidas contigo mismo es la cota superior", False),
    ("de fiabilidad de tu patron de medida, y es la respuesta a 'cuanto te fias de el'.", False),
    ("", False),
    ("No mires la salida del sistema antes de anotar. Contamina el criterio.", True),
]
for i, (t, negrita) in enumerate(texto, start=1):
    c = ins.cell(row=i, column=1, value=t)
    if negrita:
        c.font = Font(bold=True, color=ROJO if t and t[0].isupper() else "000000")
ins.column_dimensions["A"].width = 100

ruta = "datos/gold_set_PLANTILLA_v3.xlsx"
wb.save(ruta)
print("escrita:", ruta, "|", ultima - 1, "huecos de linea para 15 filas")
