"""Evalua la pasada del blind set de 300 filas.

BLOQUE A (200 filas, verdad conocida por construccion):
  se compara el ELEMENTO PRINCIPAL de cada fila contra su gold. Mide acierto sobre
  filas que el sistema no ha visto nunca, que es lo que el 0% de escape del arnes
  NO mide, porque aquel esta medido contra las mismas 15 filas de desarrollo.

BLOQUE B (100 filas adversarias):
  no hay respuesta correcta. Se mide INVENCION: que el sistema emita, en una linea
  RESUELTA, un valor de catalogo que no aparece en el texto de origen. Cero es lo
  unico aceptable: inventar cuesta 50.000 euros y dejar un hueco cuesta 1.
"""
import json
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from motor.catalogos import ACABADOS, CALIDADES_ALIAS, NOMBRES

ATRIBUTOS = ("nombre", "material", "calidad", "medida", "longitud", "norma", "acabado")
PRINCIPALES = {"TORNILLO", "ESPARRAGO", "VARILLA ROSCADA"}


def _norm(v):
    if v is None:
        return ""
    t = unicodedata.normalize("NFKC", str(v)).strip().upper()
    return " ".join(t.split())


# "N/A" y vacio significan lo mismo: el atributo no aplica a esta pieza (seccion 7
# de las reglas exime de longitud a tuercas y arandelas). Compararlos como distintos
# era un fallo del evaluador, no del sistema.
_NO_APLICA = {"", "N/A", "NA", "NO APLICA", "-"}


def _equivale(esperado, obtenido):
    if esperado in _NO_APLICA and obtenido in _NO_APLICA:
        return True
    return esperado == obtenido


def cargar_gold(ruta):
    ws = openpyxl.load_workbook(ruta)["gold"]
    gold = {}
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=1).value
        if item is None:
            continue
        gold[int(item)] = {
            "bloque": ws.cell(row=r, column=2).value,
            "descripcion": ws.cell(row=r, column=3).value or "",
            **{a: ws.cell(row=r, column=4 + i).value for i, a in enumerate(ATRIBUTOS)},
            "material_col": ws.cell(row=r, column=11).value or "",
        }
    return gold


def principal_de(lineas_fila):
    """El principal por TIPO, no por posicion. Si no hay ninguno, el primero."""
    for l in lineas_fila:
        if _norm(l["nombre"]) in PRINCIPALES:
            return l
    return lineas_fila[0] if lineas_fila else None


def _alias_de(valor, tabla):
    """Todos los literales del catalogo que normalizan a `valor`.

    Buscar el valor NORMALIZADO dentro del texto crudo marcaba como invencion
    cualquier normalizacion correcta: `ZINCADO` -> `CINCADO`, `HDG` ->
    `GALVANIZADO EN CALIENTE`. Hay que buscar los alias, no el resultado.
    """
    return [k for k, v in tabla.items() if _norm(v) == _norm(valor)] or [valor]


def evaluar(ruta_salida="datos/blind_set_salida.json",
            ruta_gold="datos/blind_set_gold.xlsx"):
    salida = json.loads(Path(ruta_salida).read_text(encoding="utf-8"))
    gold = cargar_gold(ruta_gold)

    por_fila = defaultdict(list)
    for l in salida:
        por_fila[l["fila"]].append(l)

    # ---------------- bloque A: acierto ----------------
    aciertos = Counter()
    evaluables = Counter()
    fallos = []
    filas_a = filas_a_resueltas = 0
    for item, g in gold.items():
        if g["bloque"] != "A":
            continue
        filas_a += 1
        lineas = por_fila.get(item, [])
        p = principal_de(lineas)
        if p is None:
            fallos.append((item, "SIN LINEA", g["descripcion"], "", ""))
            continue
        if p["estado"] == "RESUELTA":
            filas_a_resueltas += 1
        for a in ATRIBUTOS:
            esperado = _norm(g[a])
            if esperado == "":
                continue
            evaluables[a] += 1
            obtenido = _norm(p[a])
            if _equivale(esperado, obtenido):
                aciertos[a] += 1
            else:
                fallos.append((item, a, g["descripcion"], esperado, obtenido))

    # escape: linea RESUELTA con algun atributo distinto del gold
    escapes = set()
    for item, a, desc, esp, obt in fallos:
        p = principal_de(por_fila.get(item, []))
        if p and p["estado"] == "RESUELTA":
            escapes.add(item)

    # ---------------- bloque B: invencion ----------------
    invenciones = []
    filas_b = filas_b_resueltas = 0
    for item, g in gold.items():
        if g["bloque"] != "B":
            continue
        filas_b += 1
        # el texto de origen incluye la columna MATERIAL del xlsx, de donde el
        # sistema puede legitimamente sacar la calidad del elemento principal
        texto = _norm(g["descripcion"]) + " || " + _norm(g.get("material_col", ""))
        for l in por_fila.get(item, []):
            if l["estado"] != "RESUELTA":
                continue
            filas_b_resueltas += 1
            for a in ("calidad", "acabado", "material"):
                v = _norm(l[a])
                proc = l["procedencias"][a]
                if not v or proc in ("DERIVADO", "AUSENTE"):
                    continue
                tabla = {"acabado": ACABADOS, "calidad": CALIDADES_ALIAS}.get(a)
                literales = _alias_de(v, tabla) if tabla else [v]
                if not any(_norm(x) in texto for x in literales):
                    invenciones.append((item, a, v, proc, g["descripcion"]))

    return dict(gold=gold, por_fila=por_fila, aciertos=aciertos, evaluables=evaluables,
                fallos=fallos, escapes=escapes, filas_a=filas_a,
                filas_a_resueltas=filas_a_resueltas, invenciones=invenciones,
                filas_b=filas_b, filas_b_resueltas=filas_b_resueltas,
                lineas=len(salida))


def informe(r):
    L = []
    ev = sum(r["evaluables"].values())
    ac = sum(r["aciertos"].values())
    L.append("# Blind set de 300 filas — informe")
    L.append("")
    L.append("## Bloque A · 200 filas compuestas, verdad conocida por construcción")
    L.append("")
    L.append(f"- Filas evaluadas: {r['filas_a']}")
    L.append(f"- Resueltas: {r['filas_a_resueltas']} "
             f"({100 * r['filas_a_resueltas'] / max(1, r['filas_a']):.1f}%)")
    L.append(f"- Celdas evaluables: {ev} · aciertos: {ac} "
             f"({100 * ac / max(1, ev):.1f}%)")
    L.append(f"- **Filas RESUELTAS con algún atributo mal (escape): {len(r['escapes'])} "
             f"({100 * len(r['escapes']) / max(1, r['filas_a']):.1f}%)**")
    L.append("")
    L.append("| atributo | aciertos | evaluables | tasa |")
    L.append("|---|---|---|---|")
    for a in ATRIBUTOS:
        e, c = r["evaluables"][a], r["aciertos"][a]
        if e:
            L.append(f"| {a} | {c} | {e} | {100 * c / e:.1f}% |")
    L.append("")
    L.append("## Bloque B · 100 filas adversarias")
    L.append("")
    L.append(f"- Líneas RESUELTAS: {r['filas_b_resueltas']}")
    L.append(f"- **Invenciones (valor de catálogo que no está en el texto): "
             f"{len(r['invenciones'])}**")
    if r["invenciones"]:
        L.append("")
        for item, a, v, proc, desc in r["invenciones"][:25]:
            L.append(f"  - fila {item} · {a}={v} ({proc}) · `{desc[:70]}`")
    L.append("")
    L.append("## Primeros fallos del bloque A")
    L.append("")
    for item, a, desc, esp, obt in r["fallos"][:30]:
        L.append(f"- fila {item} · **{a}**: esperado `{esp}` · obtenido `{obt or '(vacío)'}`")
        L.append(f"  `{desc[:90]}`")
    return "\n".join(L)


if __name__ == "__main__":
    print(informe(evaluar()))
