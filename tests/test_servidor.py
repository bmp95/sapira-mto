"""Tarea 14: la API. TestClient de FastAPI + PuertoFalso, sin red -- ningun test de este
fichero llama a Gemini de verdad."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from api.servidor import crear_app
from datos.guion_falso import puerto_de_guion
from motor.modelos import Elemento, Segmentacion
from motor.puerto_llm import PuertoFalso
from motor.saneado import sanear

RUTA_MTO = Path("datos/MTO_tornilleria.xlsx")
_MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _cliente(puerto=None) -> TestClient:
    app = crear_app(puerto=puerto if puerto is not None else puerto_de_guion())
    return TestClient(app)


def _subir(cliente: TestClient, ruta: Path, nombre: str | None = None) -> dict:
    with ruta.open("rb") as f:
        resp = cliente.post(
            "/api/procesar",
            files={"archivo": (nombre or ruta.name, f, _MEDIA_XLSX)},
        )
    assert resp.status_code == 200, resp.text
    return resp.json()


def _escribir_xlsx(ruta: Path, filas: list[tuple[str, int]]) -> None:
    """Mismo layout que `motor.lectura_mto.leer_mto` espera: cabecera en las
    filas 1-4, datos desde la 5. `filas` es (descripcion, cantidad)."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    for _ in range(4):
        hoja.append([None] * 6)
    for item, (descripcion, cantidad) in enumerate(filas, start=1):
        hoja.append([item, descripcion, "", "", cantidad, "uds"])
    libro.save(ruta)


def _puerto_de_un_elemento(descripciones: list[str], tipo: str = "X") -> PuertoFalso:
    """Cada descripcion es un unico elemento que cubre la fila entera --
    mismo patron que `_escribir_mto_de_prueba`/`_PuertoQueFallaEnUnTexto` en
    tests/test_pipeline.py."""
    respuestas = {}
    for descripcion in descripciones:
        texto = sanear(descripcion)
        respuestas[texto] = Segmentacion(elementos=[Elemento(tipo_indicado=tipo, span=(0, len(texto)))])
    return PuertoFalso(respuestas=respuestas)


# --------------------------------------------------------------------------
# 1. Subir el MTO de muestra -> 30 lineas, resumen cuadra, traza completa.
# --------------------------------------------------------------------------

def test_subir_mto_de_muestra_devuelve_30_lineas_y_el_resumen_cuadra():
    cliente = _cliente()
    datos = _subir(cliente, RUTA_MTO)

    assert isinstance(datos.get("sesion_id"), str) and datos["sesion_id"]
    assert len(datos["lineas"]) == 30

    resumen = datos["resumen"]
    assert resumen["total_lineas"] == 30
    assert resumen["resueltas"] == 13
    assert resumen["en_revision"] == 17
    assert resumen["resueltas"] + resumen["en_revision"] == 30
    assert resumen["fallos_de_proceso"] == 0
    assert resumen["segundos"] >= 0
    assert resumen["coste"] >= 0

    for linea in datos["lineas"]:
        for atributo in ("nombre", "material", "calidad", "medida", "longitud", "norma", "acabado"):
            celda = linea[atributo]
            for campo in ("valor", "literal", "span", "procedencia", "regla", "factores"):
                assert campo in celda, f"{atributo}.{campo} ausente"
        for campo in ("confianza", "estado", "motivos", "cantidad", "fila_origen",
                      "texto_origen", "tramo", "id"):
            assert campo in linea, f"{campo} ausente en la linea"
        for motivo in linea["motivos"]:
            for campo in ("codigo", "texto", "atributo", "valor_propuesto", "factor_limitante"):
                assert campo in motivo


def test_el_resumen_no_imprime_ni_filtra_la_clave():
    """La clave nunca se imprime: comprueba que el resumen (y su serializacion)
    no contiene el literal GEMINI_API_KEY como si fuera a exponer un valor."""
    cliente = _cliente()
    datos = _subir(cliente, RUTA_MTO)
    assert "GEMINI_API_KEY" not in str(datos)


# --------------------------------------------------------------------------
# 2 y 3. Resolver: recalcula confianza y estado; nunca se escribe a mano.
# --------------------------------------------------------------------------

def test_resolver_una_celda_suficiente_pasa_la_linea_a_resuelta(tmp_path):
    """Tuerca con norma pero sin calidad: el unico motivo es SIN_CALIDAD.
    Al resolver la calidad, la linea llega a confianza 100 y el motivo
    desaparece."""
    ruta = tmp_path / "una_tuerca.xlsx"
    _escribir_xlsx(ruta, [("Tuerca hexagonal DIN 934 M16", 10)])
    puerto = _puerto_de_un_elemento(["Tuerca hexagonal DIN 934 M16"], tipo="TUERCA")
    cliente = _cliente(puerto)
    datos = _subir(cliente, ruta)

    assert len(datos["lineas"]) == 1
    linea = datos["lineas"][0]
    assert linea["estado"] == "REVISION_MANUAL"
    assert linea["calidad"]["procedencia"] == "AUSENTE"
    assert [m["codigo"] for m in linea["motivos"]] == ["SIN_CALIDAD"]

    resp = cliente.post("/api/resolver", json={
        "sesion_id": datos["sesion_id"], "linea_id": linea["id"],
        "atributo": "calidad", "valor": "8.8", "autor": "ana@epc.es",
    })
    assert resp.status_code == 200, resp.text
    actualizada = resp.json()

    assert actualizada["calidad"]["valor"] == "8.8"
    # No es EXTRAIDO: nadie lo leyo del MTO, lo contesto una persona. Y no
    # lleva span, porque no hay ningun trozo del texto original al que apuntar.
    assert actualizada["calidad"]["procedencia"] == "HEREDADO"
    assert actualizada["calidad"]["span"] is None
    assert "ana@epc.es" in actualizada["calidad"]["regla"]
    assert actualizada["confianza"] == 100
    assert actualizada["estado"] == "RESUELTA"
    assert [m["codigo"] for m in actualizada["motivos"]] == []


def test_resolver_una_celda_insuficiente_deja_la_linea_en_revision(tmp_path):
    """Misma tuerca, pero sin norma tampoco: SIN_NORMA y SIN_CALIDAD.
    Resolver solo la calidad no basta -- SIN_NORMA sigue ahi."""
    ruta = tmp_path / "una_tuerca_sin_norma.xlsx"
    _escribir_xlsx(ruta, [("Tuerca hexagonal M16", 10)])
    puerto = _puerto_de_un_elemento(["Tuerca hexagonal M16"], tipo="TUERCA")
    cliente = _cliente(puerto)
    datos = _subir(cliente, ruta)

    assert len(datos["lineas"]) == 1
    linea = datos["lineas"][0]
    assert linea["estado"] == "REVISION_MANUAL"
    codigos = {m["codigo"] for m in linea["motivos"]}
    assert codigos == {"SIN_NORMA", "SIN_CALIDAD"}

    resp = cliente.post("/api/resolver", json={
        "sesion_id": datos["sesion_id"], "linea_id": linea["id"],
        "atributo": "calidad", "valor": "8.8", "autor": "ana@epc.es",
    })
    assert resp.status_code == 200, resp.text
    actualizada = resp.json()

    assert actualizada["calidad"]["valor"] == "8.8"
    assert actualizada["estado"] == "REVISION_MANUAL"
    codigos_tras_resolver = {m["codigo"] for m in actualizada["motivos"]}
    assert "SIN_NORMA" in codigos_tras_resolver
    assert "SIN_CALIDAD" not in codigos_tras_resolver


def test_resolver_sesion_desconocida_da_404():
    cliente = _cliente()
    resp = cliente.post("/api/resolver", json={
        "sesion_id": "no-existe", "linea_id": "L001",
        "atributo": "calidad", "valor": "8.8", "autor": "ana@epc.es",
    })
    assert resp.status_code == 404


def test_resolver_atributo_invalido_da_400():
    cliente = _cliente()
    datos = _subir(cliente, RUTA_MTO)
    resp = cliente.post("/api/resolver", json={
        "sesion_id": datos["sesion_id"], "linea_id": datos["lineas"][0]["id"],
        "atributo": "color", "valor": "azul", "autor": "ana@epc.es",
    })
    assert resp.status_code == 400


# --------------------------------------------------------------------------
# 4. Exportar: agrupado por material canonico, cantidades sumadas.
# --------------------------------------------------------------------------

def test_exportar_agrupa_lineas_identicas_sumando_cantidad(tmp_path):
    descripcion = "Tornillo hexagonal DIN 933 M10 x 40, 8.8, zincado"
    ruta = tmp_path / "dos_iguales.xlsx"
    _escribir_xlsx(ruta, [(descripcion, 5), (descripcion, 7)])
    puerto = _puerto_de_un_elemento([descripcion], tipo="TORNILLO")
    cliente = _cliente(puerto)
    datos = _subir(cliente, ruta)

    assert len(datos["lineas"]) == 2
    assert all(l["estado"] == "RESUELTA" for l in datos["lineas"])

    resp = cliente.get("/api/exportar", params={"sesion_id": datos["sesion_id"]})
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(_MEDIA_XLSX)

    libro = openpyxl.load_workbook(BytesIO(resp.content))
    hoja = libro.active
    filas = [f for f in hoja.iter_rows(min_row=2, values_only=True) if f[0] is not None]
    assert len(filas) == 1
    cabecera = [c.value for c in hoja[1]]
    idx_cantidad = cabecera.index("cantidad")
    assert filas[0][idx_cantidad] == 12


def test_exportar_sesion_desconocida_da_404():
    cliente = _cliente()
    resp = cliente.get("/api/exportar", params={"sesion_id": "no-existe"})
    assert resp.status_code == 404


# --------------------------------------------------------------------------
# 5. Un fichero que no es xlsx: error claro, no una traza.
# --------------------------------------------------------------------------

def test_fichero_que_no_es_xlsx_da_error_claro_no_una_traza():
    cliente = _cliente()
    resp = cliente.post(
        "/api/procesar",
        files={"archivo": ("nota.txt", b"esto no es un fichero excel", "text/plain")},
    )
    assert resp.status_code == 400
    cuerpo = resp.json()
    assert "detail" in cuerpo
    assert isinstance(cuerpo["detail"], str) and len(cuerpo["detail"]) > 0
    assert "xlsx" in cuerpo["detail"].lower()


def test_resolver_sin_autor_da_422():
    """Sin quien contesta la respuesta no es auditable: no se acepta."""
    cliente = _cliente()
    datos = _subir(cliente, RUTA_MTO)
    resp = cliente.post("/api/resolver", json={
        "sesion_id": datos["sesion_id"], "linea_id": datos["lineas"][0]["id"],
        "atributo": "calidad", "valor": "8.8",
    })
    assert resp.status_code == 422


def test_lo_resuelto_a_mano_se_hereda_al_volver_a_subir_el_mto(tmp_path):
    """El ciclo entero, que es el argumento de negocio: se pregunta una vez y
    la siguiente revision de la misma pieza ya no pregunta."""
    ruta = tmp_path / "una_tuerca.xlsx"
    _escribir_xlsx(ruta, [("Tuerca hexagonal DIN 934 M16", 10)])
    puerto = _puerto_de_un_elemento(["Tuerca hexagonal DIN 934 M16"], tipo="TUERCA")
    cliente = _cliente(puerto)

    primera = _subir(cliente, ruta)
    linea = primera["lineas"][0]
    assert linea["estado"] == "REVISION_MANUAL"

    cliente.post("/api/resolver", json={
        "sesion_id": primera["sesion_id"], "linea_id": linea["id"],
        "atributo": "calidad", "valor": "8.8", "autor": "ana@epc.es",
    })

    # Misma pieza, revision siguiente: ya no hay nada que preguntar.
    segunda = _subir(cliente, ruta)
    heredada = segunda["lineas"][0]
    assert heredada["calidad"]["valor"] == "8.8"
    assert heredada["calidad"]["procedencia"] == "HEREDADO"
    assert heredada["estado"] == "RESUELTA"
    assert any(m["codigo"] == "VALOR_HEREDADO" for m in heredada["motivos"])
